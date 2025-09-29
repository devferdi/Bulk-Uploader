import tkinter as tk
from tkinter import messagebox, scrolledtext
import sys
import io
import threading
from tkinter import filedialog
import os
import pandas as pd
import concurrent.futures
from bs4 import BeautifulSoup
import json
import urllib.parse
import unicodedata
import os
import time
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
import requests
import re
from openai import OpenAI

file_lock = threading.Lock()  # 🔒 Prevents simultaneous write conflicts



def normalize_filename(filename):
    # Normalize filename to handle special characters
    return unicodedata.normalize('NFC', filename)

def file_exists_in_folder(folder, filename):
    # Normalize filename to ensure correct access to files with umlauts or other special characters
    normalized_filename = normalize_filename(filename)
    file_path = os.path.join(folder, normalized_filename)
    return os.path.exists(file_path)


# URL-encode the filename to handle special characters like umlauts
def encode_filename(filename):
    return urllib.parse.quote(filename)

# Function to convert HTML to Shopify JSON
def html_to_shopify_json(html_input):
    # Parse the HTML input using BeautifulSoup
    soup = BeautifulSoup(html_input, 'html.parser')
    
    json_structure = {"type": "root", "children": []}

    def parse_element(element):
        # Handle paragraphs
        if element.name == 'p' and element.get_text(strip=True):  # Only include if paragraph has text
            paragraph = {"type": "paragraph", "children": []}
            for child in element.children:
                if isinstance(child, str):  # Plain text
                    paragraph["children"].append({"type": "text", "value": child})
                elif child.name == 'strong':  # Bold text
                    paragraph["children"].append({"type": "text", "value": child.get_text(), "bold": True})
            print(f"Parsed paragraph: {paragraph}")  # Debug print
            return paragraph

        # Handle unordered lists
        elif element.name == 'ul':
            list_items = []
            for li in element.find_all('li'):
                list_items.append({
                    "type": "list-item",
                    "children": [{"type": "text", "value": li.get_text()}]
                })
            list_element = {"type": "list", "listType": "unordered", "children": list_items}
            print(f"Parsed unordered list: {list_element}")  # Debug print
            return list_element

    # Parse each top-level element
    for element in soup.children:
        if isinstance(element, str):  # Ignore plain text nodes
            continue
        parsed_element = parse_element(element)
        if parsed_element:
            json_structure["children"].append(parsed_element)

    # Check if json_structure contains meaningful content
    if not json_structure["children"]:
        json_structure["children"].append({
            "type": "paragraph",
            "children": [{"type": "text", "value": ""}]  # Default empty content
        })
    
    print(f"Final JSON structure: {json_structure}")  # Debug print to show final output
    return json_structure



# Function to redirect output to the GUI text box
class RedirectOutput(io.StringIO):
    def __init__(self, text_area, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.text_area = text_area

        # Ensure the redirected output uses a readable, high-contrast style. Some
        # macOS Tk builds inherit a near-white foreground, which can make the
        # text effectively invisible against the default background when the app
        # is bundled. Explicit styling guarantees the log remains legible.
        self.text_area.configure(
            bg="#FAFAFA", fg="#0F0F0F", insertbackground="#0F0F0F",
            selectbackground="#0F62FE", selectforeground="#FFFFFF"
        )
        self.text_area.tag_configure("stdout", foreground="#0F0F0F")

    def write(self, msg):
        if not msg:
            return

        def append():
            if not self.text_area.winfo_exists():
                return
            self.text_area.insert(tk.END, msg, "stdout")
            self.text_area.see(tk.END)  # Auto scroll to the end

        # Tkinter isn't thread-safe. Schedule GUI updates on the main thread.
        try:
            self.text_area.after(0, append)
        except RuntimeError:
            # If the widget is being destroyed, silently ignore further writes.
            pass

    def flush(self):
        pass  # The flush method is required for compatibility with `sys.stdout`

# Function to read store credentials from a text file
def read_credentials(file_path):
    credentials = {}
    with open(file_path, 'r') as file:
        for line in file:
            key, value = line.strip().split('=')
            credentials[key] = value
    return credentials



def run_downloader_logic():
    # Get the directory where the executable or script is located
    if getattr(sys, 'frozen', False):  # If running as an EXE
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Build the full path to 'credentials.txt'
    credentials_path = os.path.join(script_dir, 'credentials.txt')

    # Load credentials from 'credentials.txt'
    credentials = read_credentials(credentials_path)
    SHOP_NAME = credentials['store_name']
    ACCESS_TOKEN = credentials['access_token']

    # Shopify API URL
    BASE_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07"

    # Headers for API authentication
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    # The rest of your downloader logic goes here...

    # Fetch all locations
    def get_locations():
        url = f"{BASE_URL}/locations.json"
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            return response.json().get('locations', [])
        return []

    # Fetch inventory levels for a specific variant at a location
    def get_inventory_level(inventory_item_id, location_id):
        url = f"{BASE_URL}/inventory_levels.json?inventory_item_ids={inventory_item_id}&location_ids={location_id}"
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            inventory_levels = response.json().get('inventory_levels', [])
            if inventory_levels:
                return inventory_levels[0].get('available', 0)
        return None

    # Fetch products from Shopify using pagination
    def get_all_products():
        products = []
        url = f"{BASE_URL}/products.json?limit=250&fields=id,title,body_html,handle,tags,vendor,product_type,variants,images,created_at,updated_at,status,published_at,published_scope,template_suffix,options"

        print("Starting to fetch products...")
        page = 1  # Page counter

        while url:
            print(f"Fetching page {page} of products...")
            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                print(f"Error fetching products: {response.status_code}")
                break

            batch = response.json().get('products', [])
            if not batch:  # Check if the current batch is empty to avoid unnecessary processing
                print("No more products found.")
                break

            products.extend(batch)

            # Properly handle pagination if more products are available
            link_header = response.headers.get('Link')
            next_url = None
            if link_header:
                links = link_header.split(',')
                for link in links:
                    if 'rel="next"' in link:
                        next_url = link.split(';')[0].strip('<> ').replace('&amp;', '&')
                        break

            if next_url:
                url = next_url
                page += 1
            else:
                url = None

        print(f"Finished fetching products. Total products fetched: {len(products)}")
        return products




    # Fetch metafields for a specific product
    def get_metafields(owner_id, owner_resource="product"):
        print(f"Fetching metafields for {owner_resource} ID {owner_id}...")
        time.sleep(1)  # Wait for 1 second after every request


        url = f"{BASE_URL}/metafields.json?metafield[owner_id]={owner_id}&metafield[owner_resource]={owner_resource}"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            return response.json().get('metafields', [])
        else:
            print(f"Error fetching metafields for {owner_resource} ID {owner_id}: {response.status_code}")
        return []

    def get_image_url_from_gid(gid):
        query = {
            "query": f"""
            {{
                media(id: "{gid}") {{
                    ... on MediaImage {{
                        image {{
                            originalSrc
                        }}
                    }}
                }}
            }}
            """
        }
        url = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07/graphql.json"
        response = requests.post(url, json=query, headers=headers)
        if response.status_code == 200:
            data = response.json()
            return data['data']['media']['image']['originalSrc']
        else:
            print(f"Failed to fetch image URL for gid {gid}: {response.status_code}")
            return None

    def fetch_all_metafields(products):
        print("Fetching metafields for all products concurrently...")
        product_id_to_metafields = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            future_to_product = {executor.submit(get_metafields, product['id']): product for product in products}
            for future in concurrent.futures.as_completed(future_to_product):
                product = future_to_product[future]
                try:
                    metafields = future.result()
                    product_id_to_metafields[product['id']] = metafields
                except Exception as exc:
                    print(f"Product ID {product['id']} generated an exception: {exc}")
                    product_id_to_metafields[product['id']] = []
        print("Finished fetching metafields.")
        return product_id_to_metafields

    def get_inventory_levels(inventory_item_ids):
        print("Fetching inventory levels for all variants...")
        inventory_levels = []

        # Shopify allows up to 250 inventory_item_ids per request
        batch_size = 250
        for i in range(0, len(inventory_item_ids), batch_size):
            batch_inventory_item_ids = inventory_item_ids[i:i+batch_size]
            ids_str = ','.join(map(str, batch_inventory_item_ids))
            url = f"{BASE_URL}/inventory_levels.json?inventory_item_ids={ids_str}"
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                inventory_levels.extend(response.json().get('inventory_levels', []))
            else:
                print(f"Error fetching inventory levels: {response.status_code}, {response.text}")
        print("Finished fetching inventory levels.")
        return inventory_levels

    def build_inventory_level_mapping(inventory_levels):
        inventory_mapping = {}
        for level in inventory_levels:
            key = (level['inventory_item_id'], level['location_id'])
            inventory_mapping[key] = level['available']
        return inventory_mapping

    # Main function to save product, variant, and inventory data to Excel
    def save_to_excel(products, locations):
        print("Processing data and preparing to save to Excel...")
        data = []
        all_metafield_keys = set()
        max_image_columns = 0  # Track the maximum number of images

         # Collect all inventory_item_ids
        inventory_item_ids = set()
        variant_ids = []
        for product in products:
            images = product.get('images', [])
            max_image_columns = max(max_image_columns, len(images))
            for variant in product.get('variants', []):
                inventory_item_id = variant.get('inventory_item_id')
                if inventory_item_id:
                    inventory_item_ids.add(inventory_item_id)
                variant_ids.append(variant['id'])



        # Fetch inventory levels for all inventory items
        inventory_levels = get_inventory_levels(list(inventory_item_ids))
        inventory_mapping = build_inventory_level_mapping(inventory_levels)

        # Fetch metafields for all products concurrently
        product_id_to_metafields = fetch_all_metafields(products)


        # Second pass to add data, including images and inventory levels
        product_count = len(products)
        for idx, product in enumerate(products, 1):
            print(f"Processing product {idx}/{product_count}: {product['title']}")

            # Ensure that tags are correctly handled and joined as strings
            tags = product.get('tags', [])
            if isinstance(tags, str):
                tags = tags.replace(', ', ',').split(',')
            if isinstance(tags, list):
                tags = ', '.join(tags)
            else:
                tags = str(tags)

            # Build image mapping from image ID to image details (URL and alt)
            images = product.get('images', [])
            image_id_to_image = {}
            for image in images:
                image_id = image.get('id')
                if image_id:
                    image_id_to_image[image_id] = {
                        'src': image.get('src'),
                        'alt': image.get('alt', '')
                    }

            # First product row (with all product-level details)
            first_variant = product.get('variants', [])[0]  # The first variant to display in the product row
            variant_id = first_variant.get('id', '')

            # Get variant image details if available
            variant_image_url = ''
            variant_image_alt = ''
            image_id = first_variant.get('image_id')
            if image_id and image_id in image_id_to_image:
                variant_image_url = image_id_to_image[image_id]['src']
                variant_image_alt = image_id_to_image[image_id]['alt']



            product_data = {
                "Title": product['title'],
                "Handle": product['handle'],
                "ID": product['id'],
                "Body HTML": product.get('body_html', ''),
                "Vendor": product['vendor'],
                "Type": product.get('product_type', ''),
                "Tags": tags,
                "Created At": product['created_at'],
                "Updated At": product['updated_at'],
                "Status": product['status'],
                "Published": product['published_at'],
                "Published Scope": product.get('published_scope', ''),
                "Template Suffix": product.get('template_suffix', ''),
                "Variant ID": first_variant.get('id', ''),
                "Option1 Name": product.get('options', [{}])[0].get('name', '') if len(product.get('options', [])) > 0 else "",
                "Option1 Value": first_variant.get('option1', '') if len(product.get('options', [])) > 0 else "",
                "Option2 Name": product.get('options', [{}])[1].get('name', '') if len(product.get('options', [])) > 1 else "",
                "Option2 Value": first_variant.get('option2', '') if len(product.get('options', [])) > 1 else "",
                "Option3 Name": product.get('options', [{}])[2].get('name', '') if len(product.get('options', [])) > 2 else "",
                "Option3 Value": first_variant.get('option3', '') if len(product.get('options', [])) > 2 else "",
                "Variant SKU": first_variant.get('sku', ''),
                "Variant Price": first_variant.get('price', ''),  
                "Variant Compare At Price": first_variant.get('compare_at_price', ''),
                "Variant Inventory Qty": first_variant.get('inventory_quantity', 0),
                "Variant Weight": first_variant.get('weight', ''),
                "Variant Weight Unit": first_variant.get('weight_unit', ''),
                "Variant Barcode": first_variant.get('barcode', ''),
                "Continue Selling When Sold Out": first_variant.get('inventory_policy', ''),
                "Variant Image": variant_image_url,
                "Variant Image Alt": variant_image_alt,
                "Status": product['status'],  # Include status directly in the product data

            }

            # Add image URLs and alt texts dynamically after the variant data
            for i in range(max_image_columns):
                if i < len(images):
                    image_url = images[i]['src']
                    image_alt = images[i].get('alt', '')
                else:
                    image_url = None
                    image_alt = None
                product_data[f"Image {i + 1}"] = image_url
                product_data[f"Image {i + 1} Alt"] = image_alt

            # Add product-level metafields after the images
            metafields = product_id_to_metafields.get(product['id'], [])
            for metafield in metafields:
                key = metafield['key']
                value = metafield['value']
                namespace = metafield['namespace']
                field_type = metafield.get('type', 'unknown')
                column_name = f"Metafield: {namespace}.{key} [{field_type}]"
                all_metafield_keys.add(column_name)
                product_data[column_name] = value

            # Get inventory levels for each location for the first variant
            for location in locations:
                location_id = location['id']
                location_name = location['name']
                key = (first_variant.get('inventory_item_id'), location_id)
                inventory_level = inventory_mapping.get(key, 0)
                product_data[f"Inventory Available: {location_name}"] = inventory_level

            # Append the first row for the product
            data.append(product_data)

            # Additional rows for the rest of the variants (without product-level details)
            variant_count = len(product.get('variants', []))
            for v_idx, variant in enumerate(product.get('variants', [])[1:], 2):
                print(f"  Processing variant {v_idx}/{variant_count}: {variant.get('sku', '')}")
                variant_id = variant.get('id', '')
                variant_image_url = ''
                variant_image_alt = ''
                image_id = variant.get('image_id')
                if image_id and image_id in image_id_to_image:
                    variant_image_url = image_id_to_image[image_id]['src']
                    variant_image_alt = image_id_to_image[image_id]['alt']

                variant_data = {
                    "ID": "",  # Leave product ID blank for variants
                    "Variant ID": variant.get('id', ''),
                    "Option1 Name": product.get('options', [{}])[0].get('name', '') if len(product.get('options', [])) > 0 else "",
                    "Option1 Value": variant.get('option1', '') if len(product.get('options', [])) > 0 else "",
                    "Option2 Name": product.get('options', [{}])[1].get('name', '') if len(product.get('options', [])) > 1 else "",
                    "Option2 Value": variant.get('option2', '') if len(product.get('options', [])) > 1 else "",
                    "Option3 Name": product.get('options', [{}])[2].get('name', '') if len(product.get('options', [])) > 2 else "",
                    "Option3 Value": variant.get('option3', '') if len(product.get('options', [])) > 2 else "",
                    "Variant SKU": variant.get('sku', ''),
                    "Variant Price": variant.get('price', ''),
                    "Variant Compare At Price": variant.get('compare_at_price', ''),
                    "Variant Inventory Qty": variant.get('inventory_quantity', 0),
                    "Variant Weight": variant.get('weight', ''),
                    "Variant Weight Unit": variant.get('weight_unit', ''),
                    "Variant Barcode": variant.get('barcode', ''),
                    "Continue Selling When Sold Out": variant.get('inventory_policy', ''),
                    "Variant Image": variant_image_url,
                    "Variant Image Alt": variant_image_alt
                }

                # Get inventory levels for each location
                for location in locations:
                    location_id = location['id']
                    location_name = location['name']
                    key = (variant.get('inventory_item_id'), location_id)
                    inventory_level = inventory_mapping.get(key, 0)
                    variant_data[f"Inventory Available: {location_name}"] = inventory_level

                # Append the variant row
                data.append(variant_data)

            # Append a blank row for separation (optional)
            data.append({})  # Adding an empty row for grouping in Excel

        # Create DataFrame and ensure all metafield columns are present
        df = pd.DataFrame(data)
        for key in all_metafield_keys:
            if key not in df.columns:
                df[key] = None

        # Get the current date and time to append to the filename
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"shopify_products_bulk_{current_time}.xlsx"

        # Save to Excel
        df.to_excel(file_path, index=False)

        # Open the saved Excel file using openpyxl to apply freezing panes and bold styling
        wb = load_workbook(file_path)
        ws = wb.active

        # Apply bold to product rows (rows where "ID" is present)
        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[2].value:  # Assuming "ID" is in the 3rd column (0-based index)
                for cell in row:
                    cell.font = bold_font

        # Freeze the first row and first column
        ws.freeze_panes = ws['B2']  # Freeze the first row and column

        # Save the changes
        wb.save(file_path)

        print("Data has been saved successfully.")

    # Example usage
    products = get_all_products()
    locations = get_locations()
    if products and locations:
        save_to_excel(products, locations)

    print(f"Data has been saved.")

def run_uploader_logic():
    # Get the directory where the executable or script is located
    if getattr(sys, 'frozen', False):  # If running as an EXE
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Build the full path to 'credentials.txt'
    credentials_path = os.path.join(script_dir, 'credentials.txt')

    # Load credentials from 'credentials.txt'
    credentials = read_credentials(credentials_path)
    SHOP_NAME = credentials['store_name']
    ACCESS_TOKEN = credentials['access_token']

    # Shopify API URLs
    BASE_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07"
    GRAPHQL_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07/graphql.json"

    # Headers for API authentication
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }
    graphql_headers = headers.copy()

    # Image folder path
    IMAGE_FOLDER = os.path.join(script_dir, 'img')  # Adjusted to your images folder

    # Function to clean data by removing NaN values and converting them to None
    def clean_data(data):
        """
        Recursively clean data to replace NaN, inf, -inf with None, and handle invalid values for JSON serialization.
        """
        if isinstance(data, dict):
            return {k: clean_data(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [clean_data(v) for v in data]
        elif isinstance(data, float):
            # Replace NaN, inf, -inf with None
            if pd.isna(data) or data in [float('inf'), float('-inf')]:
                return None
        return data

    




    # Function to update product images
    def update_product_images(product_id, images, alt_texts):
        product_id = str(int(product_id)) if not pd.isna(product_id) else None
        if not product_id:
            print(f"Invalid product ID, skipping image update.")
            return

        # Build the image data list including alt texts
        image_data_list = []
        for image_url, alt_text in zip(images, alt_texts):
            if pd.notna(image_url) and image_url:
                image_data = {"src": image_url}
                if pd.notna(alt_text) and alt_text:
                    image_data["alt"] = alt_text
                image_data_list.append(image_data)

        if not image_data_list:
            print(f"No valid images found for product {product_id}, skipping.")
            return

        url = f"{BASE_URL}/products/{product_id}.json"
        product_data = {
            "product": {
                "id": product_id,
                "images": image_data_list,
                "alt": alt_text

            }
        }

        try:
            response = requests.put(url, headers=headers, json=product_data)
            response.raise_for_status()  # Raise an exception for HTTP errors
            print(f"Successfully updated images for product {product_id}")
        except requests.exceptions.HTTPError as err:
            print(f"HTTP error occurred while updating images for product {product_id}: {err}")
            print(f"Response: {response.text}")
        except Exception as err:
            print(f"An error occurred while updating images for product {product_id}: {err}")

    # Function to update a product on Shopify
    def update_product(product_id, updated_data):
        product_id = str(int(product_id)) if not pd.isna(product_id) else None
        if not product_id:
            print("Product ID not provided, creating a new product.")
            create_new_product(updated_data)
            return

        url = f"{BASE_URL}/products/{product_id}.json"
        updated_data["product"] = clean_data(updated_data["product"])
        response = requests.put(url, headers=headers, json=updated_data)

        if response.status_code == 404:  # Product not found, create it
            print(f"Product {product_id} not found. Creating new product.")
            create_new_product(updated_data)
        elif response.status_code == 200:
            print(f"Successfully updated product {product_id}")
        else:
            print(f"Failed to update product {product_id}: {response.status_code}, {response.text}")
        
    def update_product_by_handle(handle, updated_data):
        if not handle:
            print("Handle not provided. Cannot update or create product.")
            return

        # Fetch product by handle
        url = f"{BASE_URL}/products.json?handle={handle}"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            products = response.json().get('products', [])
            if products:
                print(f"Product with handle '{handle}' found. Updating product.")
                product_id = products[0]['id']  # Capture the product ID

                # Grab the first variant’s ID if it exists
                variants = updated_data.get("variants", [])
                if variants:
                    variant_id = variants[0].get("id")
                else:
                    variant_id = None

                url_update = f"{BASE_URL}/products/{products[0]['id']}.json"
                updated_data["product"]["handle"] = handle  # Ensure handle is included
                updated_data["product"] = clean_data(updated_data["product"])  # Clean data
                # Clean the data before sending
                updated_data = clean_data(updated_data)
                response_update = requests.put(url_update, headers=headers, json=updated_data)
                if response_update.status_code == 200:
                    print(f"Successfully updated product with handle '{handle}' and ID {product_id}.")
                    if variant_id:
                        print(f"First variant ID: {variant_id}")
                    return product_id, variant_id  # Return the product ID for further use
                else:
                    print(f"Failed to update product with handle '{handle}': {response_update.status_code}, {response_update.text}")
            else:
                print(f"Product with handle '{handle}' not found. Creating new product.")
                updated_data["product"]["handle"] = handle
                updated_data["product"] = clean_data(updated_data["product"])  # Clean data
                print("Updated Data for Product Creation:", json.dumps(updated_data, indent=4))

                # Call create_new_product and capture product_id and variant_id
                product_id, variant_id = create_new_product(updated_data)  # Capture both IDs

                # Log or use the IDs as needed
                if product_id:
                    print(f"Product created successfully with ID {product_id} and handle {handle}")
                    if variant_id:
                        print(f"First variant ID: {variant_id}")
                    else:
                        print("No variants created for this product.")
                    return product_id, variant_id  # Return both IDs

                else:
                    print("Failed to create a new product.")
                

                
        else:
            print(f"Failed to fetch product by handle '{handle}': {response.status_code}, {response.text}")

    def get_all_products_forsku():
        products = []
        url = f"{BASE_URL}/products.json?limit=250&fields=id,title,body_html,handle,tags,vendor,product_type,variants,images,created_at,updated_at,status,published_at,published_scope,template_suffix,options"

        print("Starting to fetch products...")
        page = 1  # Page counter

        while url:
            print(f"Fetching page {page} of products...")
            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                print(f"Error fetching products: {response.status_code}")
                break

            batch = response.json().get('products', [])
            if not batch:  # Check if the current batch is empty to avoid unnecessary processing
                print("No more products found.")
                break

            products.extend(batch)

            # Properly handle pagination if more products are available
            link_header = response.headers.get('Link')
            next_url = None
            if link_header:
                links = link_header.split(',')
                for link in links:
                    if 'rel="next"' in link:
                        next_url = link.split(';')[0].strip('<> ').replace('&amp;', '&')
                        break

            if next_url:
                url = next_url
                page += 1
            else:
                url = None

        print(f"Finished fetching products. Total products fetched: {len(products)}")
        return products


    def update_product_by_sku(sku, updated_data):
        if not sku:
            print("SKU not provided. Cannot update or create product.")
            return None, None
        
        print("Updated Data Received:", json.dumps(updated_data, indent=4))


        # Fetch all products
        products = get_all_products_forsku()
        found_variant = None

        # Check if any existing variant matches the SKU
        for product in products:
            for variant in product['variants']:
                if variant['sku'] == sku:
                    found_variant = variant
                    break
            if found_variant:
                break

        if found_variant:
            product_id = found_variant['product_id']  # Assuming this key exists in your data structure
            variant_id = found_variant['id']

            # Update the product with the product ID
            url_update = f"{BASE_URL}/products/{product_id}.json"
            updated_data["product"] = clean_data(updated_data["product"])  # Clean the data before sending
            response_update = requests.put(url_update, headers=headers, json=updated_data)
            if response_update.status_code == 200:
                updated_fields = ", ".join(updated_data["product"].keys())
                print(f"Successfully updated product with SKU '{sku}' and ID {product_id}. Updated fields: {updated_fields}.")
                return product_id, variant_id
            else:
                print(f"Failed to update product with SKU '{sku}': {response_update.status_code}, {response_update.text}")

        print(f"No product found with SKU '{sku}'. Attempting to create new product.")

        # If no product with the given SKU exists, create a new one
        variant_details = updated_data["product"]["variants"][0] if "variants" in updated_data["product"] and len(updated_data["product"]["variants"]) > 0 else {}
        updated_data["product"]["variants"] = [{
            'sku': sku,
            'price': variant_details.get("price"),
            'weight': variant_details.get("weight"),
            'weight_unit': variant_details.get("weight_unit", "kg")  # assuming default weight unit if not specified
        }]  # Add SKU, price, and weight to the variants in the product data
        updated_data["product"] = clean_data(updated_data["product"])  # Clean the data
        print("Updated Data for Product Creation:", json.dumps(updated_data, indent=4))
        product_id, variant_id = create_new_product(updated_data)
        if product_id:
            print(f"Product created successfully with SKU {sku} and ID {product_id}")
            if variant_id:
                print(f"First variant ID: {variant_id}")
            return product_id, variant_id
        else:
            print("Failed to create a new product.")
            return None, None


    def delete_variant(variant_id):
        if not variant_id:
            print("Variant ID not provided. Cannot delete variant.")
            return
        
        url = f"{BASE_URL}/variants/{variant_id}.json"
        response = requests.delete(url, headers=headers)
        
        if response.status_code in [200, 204]:
            print(f"Successfully deleted variant with ID {variant_id}.")
        else:
            print(f"Failed to delete variant with ID {variant_id}: {response.status_code}, {response.text}")


    def create_new_product(data):
        # Check if the necessary product information is available
        if "handle" not in data["product"] and "sku" not in data["product"].get("variants", [{}])[0]:
            print("Neither handle nor SKU provided in product data. Cannot create product.")
            return None, None  # Return None for both product ID and variant ID

        url = f"{BASE_URL}/products.json"
        response = requests.post(url, headers=headers, json=data)
        if response.status_code in [200, 201]:
            product = response.json().get("product", {})
            product_id = product.get("id")
            variants = product.get("variants", [])  # Get the variants from the response
            variant_id = variants[0].get("id") if variants else None  # Extract the first variant ID if available
            
            if product_id:
                identifier = data['product'].get('handle', data['product'].get("variants", [{}])[0].get("sku"))
                print(f"Product created successfully with ID {product_id} and identifier {identifier}")
                if variant_id:
                    print(f"First Variant ID: {variant_id}")
                return product_id, variant_id  # Return both the product ID and the first variant ID
        elif response.status_code == 422:
            print(f"Validation error while creating product: {response.text}")
        elif response.status_code == 429:
            print("Rate limit exceeded. Pausing to retry...")
            time.sleep(1)  # Delay before retrying
            return create_new_product(data)  # Retry the request
        else:
            print(f"Failed to create product: {response.status_code}, {response.text}")

        return None, None  # Return None for both product ID and variant ID if creation fails


    def update_or_create_variant(product_id, variant_id, updated_data):
        if not product_id:
            print("Product ID is required to create or update a variant.")
            return

        if variant_id:
            url = f"{BASE_URL}/variants/{variant_id}.json"
            updated_data = clean_data(updated_data)
            updated_data["variant"].pop('inventory_quantity', None)

            response = requests.put(url, headers=headers, json=updated_data)
            if response.status_code == 200:
                print(f"Successfully updated variant {variant_id}")
            elif response.status_code == 404:
                print(f"Variant {variant_id} not found. Creating new variant.")
                create_new_variant(product_id, updated_data)
            else:
                print(f"Failed to update variant {variant_id}: {response.status_code}, {response.text}")
        else:
            print("Variant ID not provided, creating a new variant.")
            create_new_variant(product_id, updated_data)


    def update_or_create_variant_by_handle(handle, variant_data):
        if not handle:
            print("Handle not provided. Cannot update or create variant.")
            return None

        # Fetch product by handle
        url = f"{BASE_URL}/products.json?handle={handle}"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            products = response.json().get('products', [])
            if products:
                product_id = products[0]['id']
                variants_url = f"{BASE_URL}/products/{product_id}/variants.json"
                variants_response = requests.get(variants_url, headers=headers)

                if variants_response.status_code == 200:
                    variants = variants_response.json().get('variants', [])
                    # Try to find a matching variant by title
                    existing_variant = next(
                        (v for v in variants if v['title'] == variant_data['variant']['option1']),
                        None
                    )
                    if existing_variant:
                        variant_id = existing_variant['id']
                        print(f"Existing variant found with ID {variant_id}. Updating variant.")
                        update_or_create_variant(product_id, variant_id, variant_data)
                        return variant_id  # Return the ID of the updated variant
                    else:
                        print("No matching variant found. Creating a new variant.")
                        new_variant_id = create_new_variant(product_id, variant_data)
                        if new_variant_id:
                            print(f"New variant created with ID {new_variant_id}.")
                        return new_variant_id  # Return the ID of the newly created variant
                else:
                    print(f"Failed to fetch variants for product ID {product_id}: {variants_response.status_code}, {variants_response.text}")
            else:
                print(f"Product with handle '{handle}' not found. Cannot create variant.")
        else:
            print(f"Failed to fetch product by handle '{handle}': {response.status_code}, {response.text}")

        return None  # Return None if no variant ID could be retrieved or created



    def create_new_variant(product_id, updated_data):
        url = f"{BASE_URL}/products/{product_id}/variants.json"

        # Ensure required fields are present
        required_options = {
            "option1": updated_data["variant"].get("option1", "Default Option1"),
            "option2": updated_data["variant"].get("option2", "Default Option2"),
            "option3": updated_data["variant"].get("option3", "")
        }
        updated_data["variant"].update(required_options)

        # Clean the data to remove NaN or unsupported values
        updated_data = clean_data(updated_data)
        updated_data["variant"]["product_id"] = product_id  # Explicitly link to product

        response = requests.post(url, headers=headers, json=updated_data)
        if response.status_code in [200, 201]:
            # Variant created successfully
            created_variant = response.json().get('variant', {})
            variant_id = created_variant.get('id')
            print(f"Variant created successfully for product ID {product_id}. Variant ID: {variant_id}")

            # Fetch all variants for the product after creating the new variant
            variants_url = f"{BASE_URL}/products/{product_id}/variants.json"
            variants_response = requests.get(variants_url, headers=headers)
            
            if variants_response.status_code == 200:
                variants = variants_response.json().get('variants', [])
                for variant in variants:
                    # Identify and delete the "Default Title" variant
                    if variant.get('title') == "Default Title":
                        print(f"Default Title variant found with ID {variant['id']}. Deleting it...")
                        delete_variant(variant['id'])
                        break  # Exit the loop after deleting the default variant
            else:
                print(f"Failed to fetch variants for product ID {product_id}: {variants_response.status_code}, {variants_response.text}")

            # Return the ID of the newly created variant
            return variant_id

        elif response.status_code == 422:
            error_message = response.json().get("errors", {})
            if "base" in error_message and "already exists" in error_message["base"][0]:
                print("Variant already exists. Skipping creation.")
            else:
                print(f"Failed to create variant: {response.status_code}, {response.text}")
        elif response.status_code == 429:
            print("Rate limit exceeded. Pausing to retry...")
            time.sleep(1)  # Delay for one second before retrying
            return create_new_variant(product_id, updated_data)
        else:
            print(f"Failed to create variant: {response.status_code}, {response.text}")

        # Return None if the creation failed
        return None






    # Function to delete a metafield
    def delete_metafield(product_id, metafield_id):
        url = f"{BASE_URL}/metafields/{metafield_id}.json"
        response = requests.delete(url, headers=headers)
        if response.status_code == 200:
            print(f"Successfully deleted metafield {metafield_id} for product {product_id}")
        else:
            print(f"Failed to delete metafield {metafield_id} for product {product_id}: {response.status_code}, {response.text}")

    # Function to create a staged upload
    def staged_upload_create(filename, mime_type):
        query = """
        mutation {
        stagedUploadsCreate(input: {
            resource: FILE,
            filename: "%s",
            mimeType: "%s",
            httpMethod: POST
        }) {
            stagedTargets {
            url
            parameters {
                name
                value
            }
            }
            userErrors {
            field
            message
            }
        }
        }
        """ % (filename, mime_type)

        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "stagedUploadsCreate" in data["data"]:
            return data["data"]["stagedUploadsCreate"]["stagedTargets"][0]
        else:
            print("Error in staged upload creation:", data)
            return None

    # Function to upload the file to the staging URL
    def upload_file_to_staging(staging_target, file_path):
        url = staging_target["url"]
        files = {"file": open(file_path, "rb")}
        form_data = {param["name"]: param["value"] for param in staging_target["parameters"]}
        response = requests.post(url, data=form_data, files=files)

        # Parsing the XML response to get the file location
        if response.status_code == 201:
            print(f"File {file_path} successfully uploaded to staging URL.")
            xml_response = ET.fromstring(response.text)
            location_url = xml_response.find('Location').text  # Extracting Location URL from XML
            return location_url
        else:
            print(f"Failed to upload file. Status: {response.status_code}, {response.text}")
            return None

    # Function to commit the file with fileCreate
    def commit_file_to_shopify(file_name, original_source):
        query = """
        mutation fileCreate($files: [FileCreateInput!]!) {
        fileCreate(files: $files) {
            files {
            id  # Fetch the gid after file creation
            alt
            createdAt
            ... on GenericFile {
                url
            }
            ... on MediaImage {
                image {
                url
                }
            }
            }
            userErrors {
            code
            field
            message
            }
        }
        }
        """
        variables = {
            "files": [
                {
                    "alt": file_name,
                    "originalSource": original_source
                }
            ]
        }
        response = requests.post(GRAPHQL_URL, json={"query": query, "variables": variables}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "fileCreate" in data["data"]:
            file_info = data["data"]["fileCreate"]["files"]
            gid = file_info[0]["id"]
            print(f"File {file_name} successfully committed to Shopify.")
            print(f"GID: {gid}")  # Print the gid here
            return file_info
        else:
            print("Error in file commit:", data)
            return None

    # Function to get image URL for a given GID
    def get_image_url_for_gid(gid):
        query = f"""
        {{
        node(id: "{gid}") {{
            ... on GenericFile {{
            url
            }}
            ... on MediaImage {{
            image {{
                url
            }}
            }}
        }}
        }}
        """
        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        if response.status_code == 200:
            data = response.json()
            node = data.get('data', {}).get('node', {})
            url = None
            if 'url' in node:
                url = node['url']
            elif 'image' in node and node['image']:
                url = node['image'].get('url')
            return url
        else:
            print(f"Failed to get image URL for GID {gid}")
            return None

   
        
    # Function to upload an image to Shopify
    def upload_image_to_shopify(file_path):
        filename = os.path.basename(file_path)

        # Normalize the filename before accessing it
        normalized_filename = normalize_filename(filename)
        folder_path = os.path.dirname(file_path)

        file_path = os.path.join(os.path.dirname(file_path), normalized_filename)

        
        # Check if file exists after normalization
        if not os.path.exists(file_path):
            print(f"Image file {normalized_filename} not found in local folder.")
            return None, None


          # Ensure the filename is URL-encoded
        encoded_filename = encode_filename(filename)    
    
    
        # Extract the file name and MIME type
        mime_type = "image/jpeg" if filename.lower().endswith(".jpg") else "image/png"

        # Step 1: Create the staged upload
        staging_target = staged_upload_create(filename, mime_type)
        if not staging_target:
            return None, None

        # Step 2: Upload the file to the staging URL
        location_url = upload_file_to_staging(staging_target, file_path)
        if not location_url:
            return None, None

        # Step 3: Commit the file to Shopify
        file_info = commit_file_to_shopify(filename, location_url)
        if file_info:
            # file_info contains the files array
            file_data = file_info[0]
            gid = file_data['id']
            # Get the URL
            url = None
            if 'url' in file_data:
                url = file_data['url']
            elif 'image' in file_data and file_data['image']:
                url = file_data['image'].get('url')
            if not url:
                # Get URL via get_image_url_for_gid
                url = get_image_url_for_gid(gid)
            
            return url, gid
        else:
            return None, None

    # Function to get all files from Shopify
    def get_all_files():
        all_files = {}
        has_next_page = True
        cursor = None

        while has_next_page:
            query = f"""
            {{
            files(first: 250{' , after: "' + cursor + '"' if cursor else ''}) {{
                edges {{
                node {{
                    id
                    alt
                    ... on GenericFile {{
                    url
                    }}
                    ... on MediaImage {{
                    image {{
                        url
                    }}
                    }}
                }}
                cursor
                }}
                pageInfo {{
                hasNextPage
                }}
            }}
            }}
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
            if response.status_code == 200:
                data = response.json()
                if "data" in data and "files" in data["data"]:
                    for file in data["data"]["files"]["edges"]:
                        node = file["node"]
                        gid = node["id"]
                        alt = node["alt"]
                        url = None
                        if 'url' in node:
                            url = node['url']
                        elif 'image' in node and node['image']:
                            url = node['image'].get('url')
                        if alt:
                            all_files[alt] = (gid, url)
                        cursor = file["cursor"]

                    has_next_page = data["data"]["files"]["pageInfo"]["hasNextPage"]
                else:
                    print("No files found or error in response.")
                    return None
            else:
                print(f"Error fetching files. Status code: {response.status_code}")
                return None

        return all_files

    # Function to update or create metafields for a product
    def update_metafields(handle, metafields, existing_files, row_index, df):
        product_id = handle
        if not product_id:
            print(f"Skipping metafield update for missing product ID.")
            return

        # Fetch current metafields to handle deletion if necessary
        current_metafields_url = f"{BASE_URL}/products/{product_id}/metafields.json"
        response = requests.get(current_metafields_url, headers=headers)
        current_metafields = response.json().get('metafields', []) if response.status_code == 200 else []
        current_metafields_dict = {f"{mf['namespace']}.{mf['key']}": mf['id'] for mf in current_metafields}

        for column, value in metafields.items():
            key_type_str = column.replace('Metafield: ', '').split(' ')
            key = key_type_str[0]
            field_type = key_type_str[1].replace('[', '').replace(']', '')

            namespace, key = key.split('.')

            # Handle deletion if the value is None
            if pd.isna(value) or value is None:
                metafield_key = f"{namespace}.{key}"
                if metafield_key in current_metafields_dict:
                    delete_metafield(product_id, current_metafields_dict[metafield_key])
                continue  # Skip to the next metafield if it's being deleted

            # For file_reference metafields
            if field_type == 'file_reference':
                if isinstance(value, str):
                    if value.startswith('gid://'):
                        # It's already a GID, use it directly
                        value_gid = value
                    elif value.startswith('http'):
                        # It's a URL, need to find the GID
                        filename = os.path.basename(value)
                        if filename in existing_files:
                            value_gid = existing_files[filename][0]
                        else:
                            value_gid = None
                    else:
                        # It's a filename
                        filename = value
                        if filename in existing_files:
                            gid, url = existing_files[filename]
                            value_gid = gid
                            # Replace cell value with GID
                            df.at[row_index, column] = value_gid
                        else:
                            # Upload the image
                            file_path_local = os.path.join(IMAGE_FOLDER, filename)
                            if os.path.exists(file_path_local):
                                url, gid = upload_image_to_shopify(file_path_local)
                                if gid:
                                    existing_files[filename] = (gid, url)
                                    value_gid = gid
                                    # Replace cell value with GiID
                                    df.at[row_index, column] = url
                                else:
                                    print(f"Failed to upload image {filename}")
                                    continue
                            else:
                                print(f"Image file {filename} not found in local folder.")
                                continue
                else:
                    # Value is not a string, skip
                    continue

                if value_gid:
                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": value_gid,
                            "type": field_type.strip()  # Strip any trailing spaces or newlines from type
                        }
                    }
                else:
                    print(f"Cannot find GID for file {filename}")
                    continue

            elif field_type == 'list.file_reference':
                if isinstance(value, str):
                    file_names = [filename.strip() for filename in value.split(",") if filename.strip()]
                    file_gids = []  # Store all GIDs for the list

                    print(f"🔍 Processing metafield '{key}' for product {product_id} with {len(file_names)} files: {file_names}")

                    for filename in file_names:
                        value_gid = None  # Reset for each file
                        
                        if filename.startswith('gid://'):
                            value_gid = filename
                            print(f"✅ Using existing GID: {value_gid}")

                        elif filename.startswith('http'):
                            base_filename = os.path.basename(filename)
                            value_gid = existing_files.get(base_filename, [None])[0]
                            print(f"🔍 Retrieved GID from existing_files for {base_filename}: {value_gid}")

                        else:
                            # It's a filename
                            if filename in existing_files:
                                gid, url = existing_files[filename]
                                value_gid = gid
                                df.at[row_index, column] = value_gid

                                print(f"✅ Found existing upload for {filename}, GID: {value_gid}")
                            else:
                                # Upload image (with compression fix)
                                file_path_local = os.path.join(IMAGE_FOLDER, filename)
                                if os.path.exists(file_path_local):
                                    print(f"📤 Uploading {filename} to Shopify...")
                                    url, gid = upload_image_to_shopify(file_path_local)  # Updated function call
                                    if gid:
                                        existing_files[filename] = (gid, url)
                                        value_gid = gid
                                        df.at[row_index, column] = url

                                        print(f"✅ Successfully uploaded {filename}, new GID: {value_gid}")
                                    else:
                                        print(f"❌ Failed to upload image {filename}")
                                        continue
                                else:
                                    print(f"⚠️ Image file {filename} not found in local folder.")
                                    continue

                        if value_gid:
                            file_gids.append(value_gid)
                        else:
                            print(f"❌ Cannot find GID for file {filename}")

                    # ✅ Debugging: Show final GIDs list before updating Shopify
                    print(f"📝 Final GID list for metafield '{key}': {file_gids}")

                    # ✅ Only update Shopify if we have at least one valid GID
                    if file_gids:
                        metafield_data = {
                            "metafield": {
                                "namespace": namespace,
                                "key": key,
                                "value":  json.dumps(file_gids),  # Store as a valid list
                                "type": "list.file_reference"  # Ensure correct metafield type
                            }
                        }
                        print(f"📡 Sending metafield update to Shopify: {metafield_data}")

                        
                    else:
                        print(f"❌ Skipping metafield update for '{key}' because the file list is empty (avoiding 422 error).")
                else:
                    print(f"⚠️ Skipping non-string value for metafield {key}.")



            else:

                if field_type == 'rich_text_field':
                    try:
                        # Ensure value is a JSON string for rich text field
                        value = json.dumps(value) if isinstance(value, dict) else value
                    except Exception as e:
                        print(f"Error serializing JSON for {namespace}.{key}: {e}")
                        continue

                # For other metafield types
                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": value,
                        "type": field_type.strip()  # Strip any trailing spaces or newlines from type
                    }
                }
                print(f"Other metafield data: {metafield_data}")


            url = f"{BASE_URL}/products/{product_id}/metafields.json"

            print(f"📡 Sending request to Shopify API: {url}")
            print(f"🔍 Headers: {json.dumps(headers, indent=2)}")
            print(f"📝 Payload: {json.dumps(metafield_data, indent=2)}")

            response = requests.post(url, headers=headers, json=metafield_data)

            print(f"📩 Shopify Response: {response.status_code}")

            if response.status_code in [200, 201]:
                print(f"✅ Successfully updated metafield {namespace}.{key} for product {product_id}")
            else:
                print(f"❌ Failed to update metafield {namespace}.{key} for product {product_id}: {response.status_code}")
                print(f"⚠️ Response Body: {response.text}")



     # Function to upload market-specific prices
    def upload_market_prices(price_list_id, market_prices):
        """
        Upload market-specific prices for product variants in bulk.
        """
        url = f"{GRAPHQL_URL}"
        mutation = """
        mutation priceListFixedPricesAdd(
          $priceListId: ID!,
          $prices: [PriceListPriceInput!]!
        ) {
          priceListFixedPricesAdd(priceListId: $priceListId, prices: $prices) {
            prices {
              variant {
                id
              }
            }
            userErrors {
              field
              message
            }
          }
        }
        """
        variables = {
            "priceListId": price_list_id,
            "prices": market_prices
        }

        response = requests.post(url, headers=graphql_headers, json={"query": mutation, "variables": variables})
        if response.status_code == 200:
            data = response.json()
            if "errors" in data:
                print(f"Errors encountered while uploading market prices: {data['errors']}")
            else:
                print(f"Successfully uploaded market prices for {len(market_prices)} variants.")
        else:
            print(f"Failed to upload market prices: {response.status_code}, {response.text}")


    # Function to upload changes from an edited spreadsheet
    def upload_changes_from_spreadsheet(file_path):
        print(f"Reading spreadsheet from: {file_path}")

        df = pd.read_excel(file_path)
        df = df.where(pd.notnull(df), None)

       

        # Get existing files mapping
        print("Fetching all existing files from Shopify...")
        existing_files = get_all_files()
        if existing_files is None:
            existing_files = {}  # To avoid errors
        print(f"Fetched {len(existing_files)} existing files from Shopify.")

        # Fetch all market names dynamically
        def get_all_market_names():
            query = """
            query Catalogs {
                catalogs(first: 10, type: MARKET) {
                    nodes {
                        ... on MarketCatalog {
                            markets(first: 10) {
                                nodes {
                                    id
                                    name
                                }
                            }
                        }
                    }
                }
            }
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=headers)
            if response.status_code == 200:
                data = response.json()
                catalogs = data.get("data", {}).get("catalogs", {}).get("nodes", [])
                market_names = []
                for catalog in catalogs:
                    markets = catalog.get("markets", {}).get("nodes", [])
                    market_names.extend([market.get("name") for market in markets if market.get("name")])
                return market_names
            print("Failed to fetch markets.")
            return []

        market_names = get_all_market_names()
        print(f"Markets found: {market_names}")

        # Identify market-specific pricing columns
        pricing_columns = {
            column: column.replace("Variant Price / ", "").strip()
            for column in df.columns
            if column.startswith("Variant Price / ")
        }
        print(f"Pricing columns identified: {pricing_columns}")

       
        def get_price_list_id_for_market(market_name):
            query = f"""
            query {{
                catalogs(first: 10, type: MARKET) {{
                    nodes {{
                        ... on MarketCatalog {{
                            markets(first: 10) {{
                                nodes {{
                                    id
                                    name
                                }}
                            }}
                            priceList {{
                                id
                            }}
                        }}
                    }}
                }}
            }}
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=headers)
            if response.status_code == 200:
                data = response.json()
                catalogs = data.get("data", {}).get("catalogs", {}).get("nodes", [])
                for catalog in catalogs:
                    markets = catalog.get("markets", {}).get("nodes", [])
                    if any(market.get("name") == market_name for market in markets):
                        price_list = catalog.get("priceList")
                        if price_list and price_list.get("id"):
                            return price_list.get("id")
            print(f"Price list ID not found for market '{market_name}'.")
            return None

        # Add helper function to add fixed prices for a market
        def add_fixed_price_for_market(price_list_id, variant_id, price_amount):
            mutation = """
            mutation priceListFixedPricesAdd($priceListId: ID!, $prices: [PriceListPriceInput!]!) {
                priceListFixedPricesAdd(priceListId: $priceListId, prices: $prices) {
                    prices {
                        variant {
                            id
                        }
                    }
                    userErrors {
                        field
                        message
                    }
                }
            }
            """
            variables = {
                "priceListId": price_list_id,
                "prices": [
                    {
                        "variantId": f"gid://shopify/ProductVariant/{variant_id}",
                        "price": {
                            "amount": str(price_amount),
                            "currencyCode": "EUR"
                        }
                    }
                ]
            }
            response = requests.post(GRAPHQL_URL, json={"query": mutation, "variables": variables}, headers=headers)
            if response.status_code == 200:
                print(f"Fixed price added for variant ID '{variant_id}' with price '{price_amount}'.")
            else:
                print(f"Failed to add fixed price for variant ID '{variant_id}': {response.text}")
        



        # Collect image filenames from the spreadsheet
        for index, row in df.iterrows():
            # Process product images
            for i in range(1, 21):  # Assuming a maximum of 20 images per product
                image_column = f"Image {i}"
                if image_column in row and row[image_column]:
                    image_value = row[image_column]
                    if isinstance(image_value, str) and not image_value.startswith(('http', 'gid://')):
                        filename = image_value
                        if filename in existing_files:
                            gid, url = existing_files[filename]
                            # Replace cell value with URL
                            df.at[index, image_column] = url
                        else:
                            # Upload image to Shopify
                            file_path_local = os.path.join(IMAGE_FOLDER, filename)
                            if os.path.exists(file_path_local):
                                with open(file_path_local, "rb") as image_file:
                                    encoded_string = base64.b64encode(image_file.read()).decode('utf-8')

                                 # Retrieve alt_text or set to default
                                alt_text = row.get("Alt Text", None)  # Use "Alt Text" or the correct column name
                                if not pd.notna(alt_text):  # If alt_text is not available or NaN
                                    alt_text = filename
                                    
                                image_data = {
                                    "image": {
                                        "attachment": encoded_string,
                                        "filename": filename,
                                        "alt": alt_text if pd.notna(alt_text) else filename
                                    }
                                }
                                product_id = row.get('ID')
                                if pd.notna(product_id):
                                    if pd.isna(product_id):
                                        product_id = None
                                    else:
                                        # Convert to string without the .0
                                        product_id = str(int(product_id))

                                    url_api = f"{BASE_URL}/products/{product_id}/images.json"
                                    response = requests.post(url_api, headers=headers, json=image_data)
                                    if response.status_code in [200, 201]:
                                        image = response.json().get('image', {})
                                        image_url = image.get('src')
                                        print(f"Successfully uploaded image {filename} to product {product_id}")
                                        # Update cell value with image URL
                                        df.at[index, image_column] = image_url
                                        existing_files[filename] = (None, image_url)
                                    else:
                                        print(f"Failed to upload image {filename} to product {product_id}: {response.status_code}, {response.text}")
                                else:
                                    print(f"Product ID missing for row {index}, cannot upload image {filename}")
                            else:
                                print(f"Image file {filename} not found in local folder.")

            


        # Initialize a variable to store the last valid handle
        last_valid_handle = None
        # Proceed with updating products and variants
        for index, row in df.iterrows():
            product_id = row.get('ID')
            variant_id = row.get('Variant ID')
            handle = row.get('Handle')  # Retrieve the handle from the spreadsheet
            sku = row.get('Variant SKU')  # Retrieve the handle from the spreadsheet

            print(f"Processing Product: Title='{sku}', Handle='{handle}'")


            if pd.isna(product_id):
                product_id = None
            else:
                # Convert to string without the .0
                product_id = str(int(product_id))

            if pd.isna(variant_id):
                variant_id = None
            else:
                # Convert to string without the .0
                variant_id = str(int(variant_id))

            handle = None if pd.isna(handle) else handle
  


            if pd.notna(handle):
                handle = handle.lower().replace(" ", "").replace("/", "-")  # Convert to lowercase, remove spaces, and replace '/' with '-'
                last_valid_handle = handle  # Update the last valid handle if the current row has one
            
                print(f"Handle='{handle}'")


            # Use SKU if handle is not available
            elif not handle and pd.notna(sku):
                print(f"No handle found. Using SKU '{sku}' as identifier.")
            elif not handle:
                print(f"Skipping row {index} due to missing handle and SKU.")
                continue

        
            options = []
            if pd.notna(handle):  # Only aggregate options for rows with a Handle
                for i in range(1, 4):  # Assuming a maximum of 3 options (Option1, Option2, Option3)
                    option_name = row.get(f"Option{i} Name")
                    if option_name and not pd.isna(option_name):  # Only add valid option names
                        # Aggregate all unique values for this option across rows with the same Handle
                        option_values = df.loc[df['Handle'] == handle, f"Option{i} Value"].dropna().unique().tolist()
                        options.append({
                            "name": option_name.strip(),
                            "values": option_values
                        })

         
            variant_name = f"{row.get('Option1 Value', '')} / {row.get('Option2 Value', '')} / {row.get('Option3 Value', '')}".strip(" /")
            if not variant_name:
                variant_name = "Default Title"

            # Skip invalid or missing Product Title and Variant Name
            if not row.get('Title') and not variant_name:
                print(f"Skipping row {index} due to missing product title and variant name.")
                continue

            # Prepare the product update data if the Product ID is available
            if row.get('Title'):

                

                # Prepare product data including options
                options = []
                # Supports up to 3 options (Option1, Option2, Option3)
                for i in range(1, 4):
                    option_name_key = f"Option{i} Name"
                    option_value_key = f"Option{i} Value"

                    # Fetch option name, set to "Title" if empty
                    option_name = row.get(option_name_key, f"Title {i}") if pd.notna(row.get(option_name_key)) else f"Title"

                    # Fetch option value, set to "Default Title" if empty
                    option_value = row.get(option_value_key, "Default Title") if pd.notna(row.get(option_value_key)) else "Default Title"

                    # Append option to the list if the option name key exists in the row (assuming that it should be there to consider)
                    if option_name_key in row:
                        options.append({
                            "name": option_name,
                            "value": option_value
                        })

                product_data = {
                    "product": {
                        "id": product_id,
                        "title": row['Title'],
                        "options": options,  # Include the options in product data
                        "variants": [  # Define variants as a list
                            {
                                "id": variant_id,
                                "price": row['Variant Price'],
                                "option1": row.get('Option1 Value', ""),  # Add Option1 Value
                                "option2": row.get('Option2 Value', ""),  # Add Option2 Value
                                "option3": row.get('Option3 Value', ""),  # Add Option3 Value
                                # Conditionally add 'sku' if present and valid
                                "sku": row.get('Variant SKU') if row.get('Variant SKU') else None,
                                # Conditionally add 'barcode' if present and valid
                                "barcode": row.get('Variant Barcode') if row.get('Variant Barcode') else None,
                                # Conditionally add 'weight' if present and valid
                                "weight": row.get('Variant Weight') if row.get('Variant Weight') else None,
                                # Conditionally add 'weight_unit' if present and valid
                                "weight_unit": row.get('Variant Weight Unit') if row.get('Variant Weight Unit') else None,
                                # Add inventory policy based on inventory quantity
                                "inventory_policy": "continue" if not pd.notna(row.get("Variant Inventory Qty")) else "deny",
                                # Add inventory quantity to update stock levels
                                "inventory_quantity": int(row["Variant Inventory Qty"]) if pd.notna(row.get("Variant Inventory Qty")) else None,
                                # Automatically set inventory management based on quantity
                                "inventory_management": "shopify" if pd.notna(row.get("Variant Inventory Qty")) else None


                            }
                        ]
                    }
                }

                # Add optional fields only if they exist or have valid values
                if row.get('Body HTML'):
                    product_data["product"]["body_html"] = row['Body HTML']


                if row.get('Type'):
                    product_data["product"]["product_type"] = row['Type']

                if row.get('Template Suffix'):
                    product_data["product"]["template_suffix"] = row['Template Suffix']

                # Conditionally add 'vendor' if present and valid
                if row.get('Vendor'):
                    product_data["product"]["vendor"] = row['Vendor']

                # Conditionally add 'tags' if present and valid
                if row.get('Tags'):
                    product_data["product"]["tags"] = row['Tags']

                if row.get('Status'):
                    product_data["product"]["status"] = row['Status']

                

            
                print(f"Updating product for handle '{handle}':", json.dumps(product_data, indent=4))
                time.sleep(1)  # Delay for half a second before retrying

                                
                if product_id and variant_id:
                    print(f"Updating product with ID '{product_id}' and variant ID '{variant_id}'...")
                    update_product(product_id, product_data)
                elif handle:
                    print(f"No valid product or variant ID found. Updating by handle '{handle}'...")
                    product_id, variant_id = update_product_by_handle(handle, product_data)
                elif sku:
                    print(f"No handle found. Updating by SKU '{sku}'...")
                    product_id, variant_id = update_product_by_sku(sku, product_data)

                if product_id:
                    print(f"Product created successfully with ID '{product_id}' and Variand ID '{variant_id}'  .")


                    # Handle market-specific pricing dynamically
                    for column, market_name in pricing_columns.items():
                        print(f"Processing column '{column}' for market '{market_name}'...")

                        if market_name in market_names:
                            print(f"Market '{market_name}' found in Shopify markets.")
                            price_amount = row.get(column)
                            print(f"Price from column '{column}': {price_amount}")

                            price_list_id = get_price_list_id_for_market(market_name)
                            print(f"Price list ID for market '{market_name}': {price_list_id}")

                            if price_list_id and price_amount:
                                print(f"Adding fixed price for market '{market_name}', variant ID '{variant_id}', price '{price_amount}'")
                                add_fixed_price_for_market(price_list_id, variant_id, price_amount)
                                print(f"Fixed price added successfully for market '{market_name}'.")
                            else:
                                if not price_list_id:
                                    print(f"[WARNING] Price list ID is missing for market '{market_name}'. Unable to add fixed price.")
                                if not price_amount:
                                    print(f"[WARNING] Price amount is missing in column '{column}' for row {index}. Skipping this market.")
                        else:
                            print(f"[INFO] Market '{market_name}' from column '{column}' not found in Shopify markets. Skipping.")


                else:
                    print("Failed to create a new product.")

                print("Updating Images and Metafields now.")


                # Collect image URLs and alt texts
                images = []
                alt_texts = []
                for i in range(1, 21):
                    image_column = f"Image {i}"
                    alt_column = f"Image {i} Alt"
                    image_url = row.get(image_column)
                    alt_text = row.get(alt_column)

                    if image_url and isinstance(image_url, str):
                        if image_url.startswith('http'):
                            # If it's already a URL, append it directly
                            images.append(image_url)
                            alt_texts.append(alt_text)
                        else:
                            # Handle local file upload (if necessary)
                            filename = image_url
                            file_path_local = os.path.join(IMAGE_FOLDER, filename)
                            if os.path.exists(file_path_local):
                                # Upload the image and get the URL
                                with open(file_path_local, "rb") as image_file:
                                    encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                                image_data = {
                                    "image": {
                                        "attachment": encoded_string,
                                        "filename": filename,
                                        "alt": alt_text if pd.notna(alt_text) else filename
                                    }
                                }
                                url_api = f"{BASE_URL}/products/{product_id}/images.json"
                                response = requests.post(url_api, headers=headers, json=image_data)
                                if response.status_code in [200, 201]:
                                    image = response.json().get('image', {})
                                    image_url = image.get('src')
                                    print(f"Successfully uploaded image {filename} to product {product_id}")
                                    # Update cell value with image URL
                                    df.at[index, image_column] = image_url
                                    images.append(image_url)
                                    alt_texts.append(alt_text)
                                else:
                                    print(f"Failed to upload image {filename} to product {product_id}: {response.status_code}, {response.text}")
                            else:
                                print(f"Image file {filename} not found in local folder.")
                    else:
                        continue  # Skip if image URL is not valid

                # Update product images with alt texts
                if images:
                    update_product_images(product_id, images, alt_texts)

                    # Retrieve the updated images to get their IDs
                    url = f"{BASE_URL}/products/{product_id}/images.json"
                    response = requests.get(url, headers=headers)
                    if response.status_code == 200:
                        product_images = response.json().get('images', [])
                        for image_data in product_images:
                            image_src = image_data.get('src')
                            image_id = image_data.get('id')
                            if image_src in images:
                                index_in_list = images.index(image_src)
                                alt_text = alt_texts[index_in_list]
                                
                    else:
                        print(f"Failed to retrieve images for product {product_id}: {response.status_code}, {response.text}")

                # Process metafields
                metafields = {}
                for column in df.columns:
                    if column.startswith('Metafield:'):
                        value = row[column]
                        # Check if it's a rich text field and contains HTML
                        if '[rich_text_field]' in column and isinstance(value, str) and '<' in value and '>' in value:
                            try:
                                json_value = html_to_shopify_json(value)  # Convert HTML to Shopify JSON format
                                metafields[column] = json_value
                            except Exception as e:
                                print(f"Error parsing HTML for {column}: {e}")
                        else:
                            # For other fields, upload the value as is
                            metafields[column] = value

                # Update metafields for the product
                if metafields:
                    if pd.notna(product_id):
                        update_metafields(product_id, metafields, existing_files, index, df)
                    else:
                        print(f"Product ID missing for row {index}, cannot update metafields.")

            # Prepare the variant update data if there is a variant
             # Prepare the variant update data if there is a variant
            if variant_name:


                variant_data = {
                "variant": {
                    "id": variant_id,
                    "price": row['Variant Price'],
                    "option1": row.get('Option1 Value', ""),  # Add Option1 Value
                    "option2": row.get('Option2 Value', ""),  # Add Option2 Value
                    "option3": row.get('Option3 Value', "")   # Add Option3 Value
                }
            }   

            # Conditionally add 'sku' if present and valid
            if row.get('Variant SKU'):
                variant_data["variant"]["sku"] = row['Variant SKU']

            # Conditionally add 'barcode' if present and valid
            if row.get('Variant Barcode'):
                variant_data["variant"]["barcode"] = row['Variant Barcode']

            # Conditionally add 'weight' if present and valid
            if row.get('Variant Weight'):
                variant_data["variant"]["weight"] = row['Variant Weight']

            # Conditionally add 'weight_unit' if present and valid
            if row.get('Variant Weight Unit'):
                variant_data["variant"]["weight_unit"] = row['Variant Weight Unit']

             # Add inventory policy based on inventory quantity
            variant_data["variant"]["inventory_policy"] = (
                "continue" if not pd.notna(row.get("Variant Inventory Qty")) else "deny"
            )


            # Add inventory quantity to update stock levels and inventory management
            if pd.notna(row.get("Variant Inventory Qty")):  # Check if the value is not NaN
                inventory_qty = int(row["Variant Inventory Qty"])
                variant_data["variant"]["inventory_quantity"] = inventory_qty
                
                # Automatically set inventory management based on the quantity
                variant_data["variant"]["inventory_management"] = "shopify" if inventory_qty > 0 else None





            print(f"Updating or creating variant for handle '{handle}':")
            time.sleep(1)  # Delay for half a second before retrying
            variant_id = update_or_create_variant_by_handle(handle, variant_data)

            if variant_id:
                print(f"Variant processed successfully with ID: {variant_id}")
                # Proceed with additional logic using `variant_id`, like setting market-specific prices

                # Handle market-specific pricing dynamically
                for column, market_name in pricing_columns.items():
                    print(f"Processing column '{column}' for market '{market_name}'...")

                    if market_name in market_names:
                        print(f"Market '{market_name}' found in Shopify markets.")
                        price_amount = row.get(column)
                        print(f"Price from column '{column}': {price_amount}")

                        price_list_id = get_price_list_id_for_market(market_name)
                        print(f"Price list ID for market '{market_name}': {price_list_id}")

                        if price_list_id and price_amount:
                            print(f"Adding fixed price for market '{market_name}', variant ID '{variant_id}', price '{price_amount}'")
                            add_fixed_price_for_market(price_list_id, variant_id, price_amount)
                            print(f"Fixed price added successfully for market '{market_name}'.")
                        else:
                            if not price_list_id:
                                print(f"[WARNING] Price list ID is missing for market '{market_name}'. Unable to add fixed price.")
                            if not price_amount:
                                print(f"[WARNING] Price amount is missing in column '{column}' for row {index}. Skipping this market.")
                    else:
                        print(f"[INFO] Market '{market_name}' from column '{column}' not found in Shopify markets. Skipping.")

            else:
                print(f"Failed to process variant for handle '{handle}'.")



        # After processing all rows, save the updated DataFrame back to Excel
        df.to_excel(file_path, index=False)
        print("Spreadsheet updated with new image URLs.")

    # Function to prompt the user to select a file
    def get_file_path():
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx")]
        )
        return file_path

    # Example usage
    file_path = get_file_path()
    if file_path:
        upload_changes_from_spreadsheet(file_path)
    else:
        print("No file selected.")# Function to run the downloader logic



def collection_run_downloader_logic():
    # Get the directory where the executable or script is located
    if getattr(sys, 'frozen', False):  # If running as an EXE
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Build the full path to 'credentials.txt'
    credentials_path = os.path.join(script_dir, 'credentials.txt')

    # Load credentials
    credentials = read_credentials(credentials_path)
    SHOP_NAME = credentials['store_name']
    ACCESS_TOKEN = credentials['access_token']

    # Shopify API URL
    BASE_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07"
    HEADERS = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    def collection_fetch_all():
        collections = []
        collection_types = ["smart_collections", "custom_collections"]
        
        for collection_type in collection_types:
            url = f"{BASE_URL}/{collection_type}.json?limit=250"
            while url:
                response = requests.get(url, headers=HEADERS)
                if response.status_code != 200:
                    print(f"Error fetching {collection_type}: {response.status_code}")
                    break
                
                data = response.json().get(collection_type, [])
                collections.extend(data)
                
                # Handle pagination
                link_header = response.headers.get("Link")
                next_url = None
                if link_header:
                    links = link_header.split(',')
                    for link in links:
                        if 'rel="next"' in link:
                            next_url = link.split(';')[0].strip('<> ')
                            break
                url = next_url
                time.sleep(1)
        
        return collections

    def collection_fetch_metafields(collection_id):
        metafields = []
        url = f"{BASE_URL}/collections/{collection_id}/metafields.json?limit=250"
        
        response = requests.get(url, headers=HEADERS)
        if response.status_code == 200:
            metafields = response.json().get("metafields", [])
        else:
            print(f"Error fetching metafields for collection {collection_id}: {response.status_code}")
        
        return metafields

    def collection_fetch_products(collection_id):
        products = []
        url = f"{BASE_URL}/collects.json?collection_id={collection_id}&limit=250"
        while url:
            response = requests.get(url, headers=HEADERS)
            if response.status_code != 200:
                print(f"Error fetching products for collection {collection_id}: {response.status_code}")
                break
            
            data = response.json().get("collects", [])
            products.extend([item["product_id"] for item in data])
            
            link_header = response.headers.get("Link")
            next_url = None
            if link_header:
                links = link_header.split(',')
                for link in links:
                    if 'rel="next"' in link:
                        next_url = link.split(';')[0].strip('<> ')
                        break
            url = next_url
            time.sleep(1)
        
        return products

    def collection_save_to_excel(collections, metafields_data, collection_products):
        print("Processing data and preparing to save to Excel...")
        data = []
        all_metafield_keys = set()
        
        for collection in collections:
            print(f"Processing collection: {collection['title']}")
            collection_type = "Smart" if "rules" in collection else "Manual"
            metafields = [mf for mf in metafields_data if mf[0] == collection["id"]]
            products = [cp[1] for cp in collection_products if cp[0] == collection["id"]]
            
            rules = collection.get("rules", [])
            rule_conditions = ", ".join([f"{rule['column']} {rule['relation']} {rule['condition']}" for rule in rules])
            
            collection_data = {
                "ID": collection["id"],
                "Title": collection["title"],
                "Handle": collection["handle"],
                "Collection Type": collection_type,
                "Created At": collection.get("created_at", ""),
                "Updated At": collection.get("updated_at", ""),
                "Image Src": collection.get("image", {}).get("src", ""),
                "Products": ", ".join(map(str, products)),
                "Conditions": rule_conditions
            }
            
            for mf in metafields:
                column_name = f"Metafield: {mf[1]}.{mf[2]} [{mf[4]}]"
                all_metafield_keys.add(column_name)
                collection_data[column_name] = mf[3]
            
            data.append(collection_data)
        
        df = pd.DataFrame(data)
        for key in all_metafield_keys:
            if key not in df.columns:
                df[key] = None
                
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"shopify_collections_{current_time}.xlsx"

        df.to_excel(file_path, index=False)
        
        wb = load_workbook(file_path)
        ws = wb.active
        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value:
                for cell in row:
                    cell.font = bold_font
        ws.freeze_panes = ws['B2']
        wb.save(file_path)
        
        print("Data has been saved successfully.")

    print("Fetching collections...")
    collections = collection_fetch_all()
    
    print("Fetching metafields and products for each collection...")
    metafields_data = []
    collection_products = []
    
    for col in collections:
        col_id = col["id"]
        metafields = collection_fetch_metafields(col_id)
        for mf in metafields:
            metafields_data.append([col_id, mf["namespace"], mf["key"], mf["value"], mf.get("type", "")])
        
        products = collection_fetch_products(col_id)
        for prod in products:
            collection_products.append([col_id, prod])
    
    print("Saving data to Excel...")
    collection_save_to_excel(collections, metafields_data, collection_products)
    print("Collection download completed.")

def collection_run_uploader_logic():
    # Get script directory
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Load credentials
    credentials_path = os.path.join(script_dir, 'credentials.txt')
    credentials = read_credentials(credentials_path)
    SHOP_NAME = credentials['store_name']
    ACCESS_TOKEN = credentials['access_token']

    # Shopify API URLs
    BASE_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07"
    GRAPHQL_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07/graphql.json"

    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    graphql_headers = headers.copy()

    IMAGE_FOLDER = os.path.join(script_dir, 'img')  # Adjust as needed
    
    def create_collection(collection_data, is_smart=False):
        url = f"{BASE_URL}/smart_collections.json" if is_smart else f"{BASE_URL}/custom_collections.json"
        print(f"📡 Creating {'Smart' if is_smart else 'Manual'} Collection: {collection_data}")
        response = requests.post(url, headers=headers, json=collection_data)
        
        print(f"📩 Response: {response.status_code} - {response.text}")

        if response.status_code in [200, 201]:
            return response.json().get("smart_collection" if is_smart else "custom_collection", {})
        else:
            print(f"❌ Failed to create collection: {response.status_code}, {response.text}")
            return None

    def delete_metafield(collection_id, metafield_id):
        """Deletes a metafield from a collection."""
        url = f"{BASE_URL}/metafields/{metafield_id}.json"
        print(f"🗑 Deleting metafield {metafield_id} from collection {collection_id}...")
        response = requests.delete(url, headers=headers)
        
        print(f"📩 Delete Response: {response.status_code}")

        if response.status_code == 200:
            print(f"✅ Deleted metafield {metafield_id} for collection {collection_id}")
        else:
            print(f"❌ Failed to delete metafield {metafield_id}: {response.status_code}, {response.text}")

    # Function to create a staged upload
    def staged_upload_create(filename, mime_type):
        query = """
        mutation {
        stagedUploadsCreate(input: {
            resource: FILE,
            filename: "%s",
            mimeType: "%s",
            httpMethod: POST
        }) {
            stagedTargets {
            url
            parameters {
                name
                value
            }
            }
            userErrors {
            field
            message
            }
        }
        }
        """ % (filename, mime_type)

        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "stagedUploadsCreate" in data["data"]:
            return data["data"]["stagedUploadsCreate"]["stagedTargets"][0]
        else:
            print("Error in staged upload creation:", data)
            return None

    # Function to upload the file to the staging URL
    def upload_file_to_staging(staging_target, file_path):
        url = staging_target["url"]
        files = {"file": open(file_path, "rb")}
        form_data = {param["name"]: param["value"] for param in staging_target["parameters"]}
        response = requests.post(url, data=form_data, files=files)

        # Parsing the XML response to get the file location
        if response.status_code == 201:
            print(f"File {file_path} successfully uploaded to staging URL.")
            xml_response = ET.fromstring(response.text)
            location_url = xml_response.find('Location').text  # Extracting Location URL from XML
            return location_url
        else:
            print(f"Failed to upload file. Status: {response.status_code}, {response.text}")
            return None

    # Function to commit the file with fileCreate
    def commit_file_to_shopify(file_name, original_source):
        query = """
        mutation fileCreate($files: [FileCreateInput!]!) {
        fileCreate(files: $files) {
            files {
            id  # Fetch the gid after file creation
            alt
            createdAt
            ... on GenericFile {
                url
            }
            ... on MediaImage {
                image {
                url
                }
            }
            }
            userErrors {
            code
            field
            message
            }
        }
        }
        """
        variables = {
            "files": [
                {
                    "alt": file_name,
                    "originalSource": original_source
                }
            ]
        }
        response = requests.post(GRAPHQL_URL, json={"query": query, "variables": variables}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "fileCreate" in data["data"]:
            file_info = data["data"]["fileCreate"]["files"]
            gid = file_info[0]["id"]
            print(f"File {file_name} successfully committed to Shopify.")
            print(f"GID: {gid}")  # Print the gid here
            return file_info
        else:
            print("Error in file commit:", data)
            return None

    # Function to get image URL for a given GID
    def get_image_url_for_gid(gid):
        query = f"""
        {{
        node(id: "{gid}") {{
            ... on GenericFile {{
            url
            }}
            ... on MediaImage {{
            image {{
                url
            }}
            }}
        }}
        }}
        """
        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        if response.status_code == 200:
            data = response.json()
            node = data.get('data', {}).get('node', {})
            url = None
            if 'url' in node:
                url = node['url']
            elif 'image' in node and node['image']:
                url = node['image'].get('url')
            return url
        else:
            print(f"Failed to get image URL for GID {gid}")
            return None

    def upload_image_to_shopify(file_path):
        filename = os.path.basename(file_path)

        # Normalize the filename before accessing it
        normalized_filename = normalize_filename(filename)
        folder_path = os.path.dirname(file_path)

        file_path = os.path.join(os.path.dirname(file_path), normalized_filename)

        
        # Check if file exists after normalization
        if not os.path.exists(file_path):
            print(f"Image file {normalized_filename} not found in local folder.")
            return None, None


          # Ensure the filename is URL-encoded
        encoded_filename = encode_filename(filename)    
    
    
        # Extract the file name and MIME type
        mime_type = "image/jpeg" if filename.lower().endswith(".jpg") else "image/png"

        # Step 1: Create the staged upload
        staging_target = staged_upload_create(filename, mime_type)
        if not staging_target:
            return None, None

        # Step 2: Upload the file to the staging URL
        location_url = upload_file_to_staging(staging_target, file_path)
        if not location_url:
            return None, None

        # Step 3: Commit the file to Shopify
        file_info = commit_file_to_shopify(filename, location_url)
        if file_info:
            # file_info contains the files array
            file_data = file_info[0]
            gid = file_data['id']
            # Get the URL
            url = None
            if 'url' in file_data:
                url = file_data['url']
            elif 'image' in file_data and file_data['image']:
                url = file_data['image'].get('url')
            if not url:
                # Get URL via get_image_url_for_gid
                url = get_image_url_for_gid(gid)
                
            print(f"✅ Image uploaded successfully: GID={gid}, URL={url}")

            return url, gid
        else:
            return None, None

    def get_all_files():
        """Retrieves all existing file references in Shopify."""
        print("📡 Fetching all existing Shopify files...")
        all_files = {}
        has_next_page = True
        cursor = None

        while has_next_page:
            query = f"""
            {{
            files(first: 250{' , after: "' + cursor + '"' if cursor else ''}) {{
                edges {{
                node {{
                    id
                    alt
                    ... on GenericFile {{
                    url
                    }}
                    ... on MediaImage {{
                    image {{
                        url
                    }}
                    }}
                }}
                cursor
                }}
                pageInfo {{
                hasNextPage
                }}
            }}
            }}
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=headers)
            if response.status_code == 200:
                data = response.json()
                for file in data.get("data", {}).get("files", {}).get("edges", []):
                    node = file["node"]
                    gid = node["id"]
                    alt = node["alt"]
                    url = node.get('url') or node.get('image', {}).get('url')
                    if alt:
                        all_files[alt] = (gid, url)
                    cursor = file["cursor"]

                has_next_page = data["data"]["files"]["pageInfo"]["hasNextPage"]
                print(f"✅ {len(all_files)} files retrieved so far...")
            else:
                print(f"❌ Error fetching files. Response: {response.status_code}")
                return None

        print(f"✅ Finished fetching {len(all_files)} existing files.")
        return all_files

    def update_metafields(collection_id, metafields, existing_files, row_index, df):
        """Handles metafield updates for collections, including file uploads and deletions."""
        
        if not collection_id:
            print(f"⚠️ Skipping metafield update: Missing collection ID.")
            return

        print(f"\n📡 Fetching existing metafields for Collection ID: {collection_id}...")
        
        # Fetch current metafields
        url = f"{BASE_URL}/collections/{collection_id}/metafields.json"
        response = requests.get(url, headers=headers)
        current_metafields = response.json().get('metafields', []) if response.status_code == 200 else []
        current_metafields_dict = {f"{mf['namespace']}.{mf['key']}": mf['id'] for mf in current_metafields}

        print(f"🔍 Found {len(current_metafields)} existing metafields.")

        for column, value in metafields.items():
            key_type_str = column.replace('Metafield: ', '').split(' ')
            key = key_type_str[0]
            field_type = key_type_str[1].replace('[', '').replace(']', '')

            namespace, key = key.split('.')

            print(f"\n📝 Processing metafield → Namespace: '{namespace}', Key: '{key}', Type: '{field_type}', Value: {value}")

            # ✅ Handle deletion if the value is None
            if pd.isna(value) or value is None:
                metafield_key = f"{namespace}.{key}"
                if metafield_key in current_metafields_dict:
                    print(f"🗑️ Deleting metafield '{namespace}.{key}' (ID: {current_metafields_dict[metafield_key]}) from collection {collection_id}...")
                    delete_metafield(collection_id, current_metafields_dict[metafield_key])
                else:
                    print(f"⚠️ Metafield '{namespace}.{key}' does not exist, skipping deletion.")
                continue

            metafield_data = None

            # ✅ Handle file_reference metafields
            if field_type == 'file_reference':
                value_gid = None
                if isinstance(value, str):
                    if value.startswith('gid://'):
                        print(f"✅ Using existing GID: {value}")
                        value_gid = value
                    elif value.startswith('http'):
                        filename = os.path.basename(value)
                        value_gid = existing_files.get(filename, [None])[0]
                        print(f"🔍 Retrieved GID from existing files: {value_gid}")
                    else:
                        filename = value
                        if filename in existing_files:
                            gid, url = existing_files[filename]
                            value_gid = gid
                            df.at[row_index, column] = value_gid
                            print(f"✅ Found existing upload: {filename}, GID: {value_gid}")
                        else:
                            file_path_local = os.path.join(IMAGE_FOLDER, filename)
                            if os.path.exists(file_path_local):
                                print(f"📤 Uploading {filename} to Shopify...")
                                url, gid = upload_image_to_shopify(file_path_local)
                                if gid:
                                    existing_files[filename] = (gid, url)
                                    value_gid = gid
                                    df.at[row_index, column] = gid
                                    print(f"✅ Successfully uploaded '{filename}', New GID: {value_gid}")
                                else:
                                    print(f"❌ Failed to upload image '{filename}', skipping metafield update.")
                                    continue
                            else:
                                print(f"⚠️ Image file '{filename}' not found in local folder, skipping metafield update.")
                                continue

                if value_gid:
                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": value_gid,
                            "type": "file_reference"
                        }
                    }

            elif field_type == 'list.file_reference':
                if isinstance(value, str):
                    file_names = [filename.strip() for filename in value.split(",") if filename.strip()]
                    file_gids = []  # Store all GIDs for the list

                    print(f"🔍 Processing metafield '{key}'  with {len(file_names)} files: {file_names}")

                    for filename in file_names:
                        value_gid = None  # Reset for each file
                        
                        if filename.startswith('gid://'):
                            value_gid = filename
                            print(f"✅ Using existing GID: {value_gid}")

                        elif filename.startswith('http'):
                            base_filename = os.path.basename(filename)
                            value_gid = existing_files.get(base_filename, [None])[0]
                            print(f"🔍 Retrieved GID from existing_files for {base_filename}: {value_gid}")

                        else:
                            # It's a filename
                            if filename in existing_files:
                                gid, url = existing_files[filename]
                                value_gid = gid
                                df.at[row_index, column] = value_gid

                                print(f"✅ Found existing upload for {filename}, GID: {value_gid}")
                            else:
                                # Upload image (with compression fix)
                                file_path_local = os.path.join(IMAGE_FOLDER, filename)
                                if os.path.exists(file_path_local):
                                    print(f"📤 Uploading {filename} to Shopify...")
                                    url, gid = upload_image_to_shopify(file_path_local)  # Updated function call
                                    if gid:
                                        existing_files[filename] = (gid, url)
                                        value_gid = gid
                                        df.at[row_index, column] = url

                                        print(f"✅ Successfully uploaded {filename}, new GID: {value_gid}")
                                    else:
                                        print(f"❌ Failed to upload image {filename}")
                                        continue
                                else:
                                    print(f"⚠️ Image file {filename} not found in local folder.")
                                    continue

                        if value_gid:
                            file_gids.append(value_gid)
                        else:
                            print(f"❌ Cannot find GID for file {filename}")

                    # ✅ Debugging: Show final GIDs list before updating Shopify
                    print(f"📝 Final GID list for metafield '{key}': {file_gids}")

                    # ✅ Only update Shopify if we have at least one valid GID
                    if file_gids:
                        metafield_data = {
                            "metafield": {
                                "namespace": namespace,
                                "key": key,
                                "value":  json.dumps(file_gids),  # Store as a valid list
                                "type": "list.file_reference"  # Ensure correct metafield type
                            }
                        }
                        print(f"📡 Sending metafield update to Shopify: {metafield_data}")

                        
                    else:
                        print(f"❌ Skipping metafield update for '{key}' because the file list is empty (avoiding 422 error).")
                else:
                    print(f"⚠️ Skipping non-string value for metafield {key}.")

            # ✅ Handle single-line text fields
            elif field_type == 'single_line_text_field':
                print(f"📌 Adding single-line text metafield: '{value}'")
                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": str(int(value)) if isinstance(value, float) and value.is_integer() else str(value),
                        "type": "single_line_text_field"
                    }
                }

            # ✅ Handle multi-line text fields (Fix for your issue)
            elif field_type == 'multi_line_text_field':
                print(f"📌 Adding multi-line text metafield: '{value}'")
                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": str(int(value)) if isinstance(value, float) and value.is_integer() else str(value),
                        "type": "multi_line_text_field"
                    }
                }

            # ✅ Send metafield update request
            if metafield_data:
                url = f"{BASE_URL}/collections/{collection_id}/metafields.json"
                print(f"📡 Sending metafield update to Shopify API: {url}")
                print(f"📝 Payload: {json.dumps(metafield_data, indent=2, ensure_ascii=False)}")  # Ensure UTF-8 support

                response = requests.post(url, headers=headers, json=metafield_data)

                if response.status_code in [200, 201]:
                    print(f"✅ Successfully updated metafield '{namespace}.{key}' for collection {collection_id}")
                else:
                    print(f"❌ Failed to update metafield '{namespace}.{key}' for collection {collection_id}: {response.status_code}")
                    print(f"⚠️ Response Body: {response.text}")

    def upload_collections_from_file(file_path):
        print(f"📂 Reading collections from file: {file_path}")
        df = pd.read_excel(file_path)
        df = df.where(pd.notnull(df), None)
        existing_files = get_all_files() or {}

        for _, row in df.iterrows():
            title = row["Title"]
            handle = row["Handle"]
            published = str(row.get("Published", "yes")).strip().lower() == "yes"

            # ✅ Extract conditions for smart collections
            conditions = []
            if "Conditions" in df.columns and pd.notna(row["Conditions"]):
                condition_strings = str(row["Conditions"]).split(";")
                for condition in condition_strings:
                    parts = condition.strip().split()
                    if len(parts) >= 3:
                        column, relation, condition_value = parts[0], parts[1], " ".join(parts[2:])
                        conditions.append({
                            "column": column,
                            "relation": relation,
                            "condition": condition_value
                        })

            is_smart = bool(conditions)

            print(f"🔍 Checking for existing collection with handle: {handle}")
            existing = find_existing_collection_by_handle(handle)

            if existing:
                collection_id = existing['id']
                print(f"🔄 Updating existing collection with ID {collection_id}")

                if is_smart:
                    collection_data = {
                        "smart_collection": {
                            "id": collection_id,
                            "title": title,
                            "handle": handle,
                            "published": published,
                            "rules": conditions,
                            "disjunctive": False
                        }
                    }
                else:
                    collection_data = {
                        "custom_collection": {
                            "id": collection_id,
                            "title": title,
                            "handle": handle,
                            "published": published
                        }
                    }

                success = update_collection(collection_id, collection_data, is_smart)
                if success:
                    print(f"✅ Updated Collection: {title} (ID: {collection_id})")
                    metafields = {col: row[col] for col in df.columns if col.startswith("Metafield:")}
                    if metafields:
                        update_metafields(collection_id, metafields, existing_files, _, df)
                else:
                    print(f"❌ Failed to update collection: {title}")
                continue  # Skip creation

            # ✅ If no existing collection, proceed to create a new one
            print(f"🆕 Creating {'Smart' if is_smart else 'Manual'} Collection: Title={title}, Handle={handle}, Published={published}")
            print(f"🛠 Conditions: {conditions if is_smart else 'None (Manual Collection)'}")

            if is_smart:
                collection_data = {
                    "smart_collection": {
                        "title": title,
                        "handle": handle,
                        "published": published,
                        "rules": conditions,
                        "disjunctive": False
                    }
                }
            else:
                collection_data = {
                    "custom_collection": {
                        "title": title,
                        "handle": handle,
                        "published": published
                    }
                }

            collection = create_collection(collection_data, is_smart)
            if collection:
                collection_id = collection.get("id")
                print(f"✅ Created Collection: {title} (ID: {collection_id})")
                metafields = {col: row[col] for col in df.columns if col.startswith("Metafield:")}
                if metafields:
                    update_metafields(collection_id, metafields, existing_files, _, df)
            else:
                print(f"❌ Failed to create collection: {title}")

    print("🎉 Collections upload completed.")

    def update_collection(collection_id, collection_data, is_smart=False):
        url = f"{BASE_URL}/smart_collections/{collection_id}.json" if is_smart else f"{BASE_URL}/custom_collections/{collection_id}.json"
        print(f"🛠 Updating {'Smart' if is_smart else 'Manual'} Collection ID {collection_id}")
        response = requests.put(url, headers=headers, json=collection_data)
        print(f"📩 Update Response: {response.status_code} - {response.text}")
        return response.status_code in [200, 201]

    def find_existing_collection_by_handle(handle):
        """Returns collection object (manual or smart) by handle, or None."""
        url = f"{BASE_URL}/custom_collections.json?handle={handle}"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            collections = response.json().get('custom_collections', [])
            if collections:
                return collections[0]
        
        # If not found in manual, try smart
        url = f"{BASE_URL}/smart_collections.json?handle={handle}"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            collections = response.json().get('smart_collections', [])
            if collections:
                return collections[0]
        
        return None


    def get_file_path():
        root = tk.Tk()
        root.withdraw()
        return filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])

    file_path = get_file_path()
    if file_path:
        upload_collections_from_file(file_path)
    else:
        print("❌ No file selected.")

def start_download():
    def after_download():
        download_button.config(state=tk.NORMAL)
        upload_button.config(state=tk.NORMAL)
        messagebox.showinfo("Success", "Download completed!")

    download_button.config(state=tk.DISABLED)
    upload_button.config(state=tk.DISABLED)

    # Start download in a separate thread
    thread = threading.Thread(target=run_downloader_logic)
    thread.start()

    # Wait for the thread to finish and then re-enable buttons
    root.after(100, check_thread, thread, after_download)

def start_upload():
    def after_upload():
        download_button.config(state=tk.NORMAL)
        upload_button.config(state=tk.NORMAL)
        messagebox.showinfo("Success", "Upload completed!")

    download_button.config(state=tk.DISABLED)
    upload_button.config(state=tk.DISABLED)

    # Start upload in a separate thread
    thread = threading.Thread(target=run_uploader_logic)
    thread.start()

    # Wait for the thread to finish and then re-enable buttons
    root.after(100, check_thread, thread, after_upload)

def start_collection_download():
    def after_collection_download():
        collection_download_button.config(state=tk.NORMAL)
        collection_upload_button.config(state=tk.NORMAL)
        messagebox.showinfo("Success", "Download completed!")
    
    collection_download_button.config(state=tk.DISABLED)
    collection_upload_button.config(state=tk.DISABLED)
    
    thread = threading.Thread(target=collection_run_downloader_logic)
    thread.start()
    root.after(100, check_thread, thread, after_collection_download)

def start_collection_upload():
    def after_collection_upload():
        collection_download_button.config(state=tk.NORMAL)
        collection_upload_button.config(state=tk.NORMAL)
        messagebox.showinfo("Success", "Upload completed!")
    
    collection_download_button.config(state=tk.DISABLED)
    collection_upload_button.config(state=tk.DISABLED)
    
    thread = threading.Thread(target=collection_run_uploader_logic)
    thread.start()
    root.after(100, check_thread, thread, after_collection_upload)

def download_shopify_files_alt_texts():
    import requests
    import pandas as pd
    import os
    from datetime import datetime

    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    
    credentials = read_credentials(os.path.join(script_dir, 'credentials.txt'))
    SHOP_NAME = credentials['store_name']
    ACCESS_TOKEN = credentials['access_token']

    GRAPHQL_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07/graphql.json"
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    print("📥 Fetching all uploaded Shopify files...")
    all_files = []
    has_next_page = True
    cursor = None

    while has_next_page:
        query = f"""
        {{
          files(first: 250 {', after: "' + cursor + '" ' if cursor else ''}) {{
            edges {{
              node {{
                id
                alt
                __typename
                ... on GenericFile {{
                  url
                }}
                ... on MediaImage {{
                  image {{
                    url
                  }}
                }}
              }}
              cursor
            }}
            pageInfo {{
              hasNextPage
            }}
          }}
        }}
        """

        response = requests.post(GRAPHQL_URL, headers=headers, json={"query": query})
        data = response.json()

        if "errors" in data:
            print(f"❌ Error fetching files: {data['errors']}")
            break

        files = data['data']['files']['edges']
        for file in files:
            node = file['node']
            gid = node['id']
            alt = node.get('alt', '')
            url = node.get('url') or (node.get('image', {}).get('url') if node.get('image') else None)
            filename = os.path.basename(url) if url else ""

            all_files.append({
                "Filename": filename,
                "Alt Text": alt,
                "GID": gid,
                "URL": url
            })

        has_next_page = data['data']['files']['pageInfo']['hasNextPage']
        if has_next_page:
            cursor = files[-1]['cursor']
    
    # Save to Excel
    df = pd.DataFrame(all_files)
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"shopify_uploaded_files_alt_texts_{current_time}.xlsx"
    df.to_excel(file_name, index=False)

    print(f"✅ Files Alt Texts exported to {file_name}")

def upload_shopify_files_alt_texts():
    import requests
    import pandas as pd
    import os
    import tkinter as tk
    from tkinter import filedialog

    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    credentials = read_credentials(os.path.join(script_dir, 'credentials.txt'))
    SHOP_NAME = credentials['store_name']
    ACCESS_TOKEN = credentials['access_token']

    GRAPHQL_URL = f"https://{SHOP_NAME}.myshopify.com/admin/api/2023-07/graphql.json"
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Shopify Files Alt Text Excel File",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        print("❌ No file selected.")
        return

    df = pd.read_excel(file_path)
        
    # 🛠 FIX: Replace NaN with empty string to avoid JSON errors
    df = df.fillna('')

    if "GID" not in df.columns or "New Alt Text" not in df.columns:
        print("❌ Excel must have 'GID' and 'New Alt Text' columns.")
        return

    for idx, row in df.iterrows():
        gid = row['GID']
        new_alt_text = row['New Alt Text']

        if not gid or not new_alt_text:
            print(f"⚠️ Skipping row {idx} due to missing GID or New Alt Text.")
            continue

        mutation = """
        mutation fileUpdate($id: ID!, $alt: String!) {
          fileUpdate(files: {id: $id, alt: $alt}) {
            files {
              id
              alt
            }
            userErrors {
              field
              message
            }
          }
        }
        """

        variables = {
            "id": gid,
            "alt": new_alt_text
        }

        response = requests.post(GRAPHQL_URL, headers=headers, json={"query": mutation, "variables": variables})
        result = response.json()

        if "errors" in result:
            print(f"❌ Error updating file {gid}: {result['errors']}")
        elif result['data']['fileUpdate']['userErrors']:
            print(f"❌ Shopify User Errors for file {gid}: {result['data']['fileUpdate']['userErrors']}")
        else:
            print(f"✅ Successfully updated Alt Text for File {gid}")

    print("✅ All Alt Text updates completed.")

def generate_seo_alt_texts():
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    credentials = read_credentials(os.path.join(script_dir, 'credentials.txt'))
    openai_api_key = credentials['openai_api_key']
    store_name = credentials['store_name']

    client = OpenAI(api_key=openai_api_key)

    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Excel File with Image URLs",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        print("❌ No file selected.")
        return

    df = pd.read_excel(file_path)
    if "URL" not in df.columns or "Filename" not in df.columns or "Alt Text" not in df.columns:
        print("❌ Excel must have 'Filename', 'Alt Text', and 'URL' columns.")
        return

    df["New Alt Text"] = ""

    print(f"🔍 Processing {len(df)} images...")

    store_name = beautify_store_name(credentials['store_name'])

    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_output_file = f"shopify_images_with_new_alt_texts_{current_time}.xlsx"
    temp_file = f"temp_alt_texts_{current_time}.xlsx"
    backup_interval = 5

    for index, row in df.iterrows():
        image_url = row['URL']
        filename = row['Filename']
        old_alt_text = row['Alt Text']

        if not image_url or not isinstance(image_url, str):
            print(f"⚠️ Skipping row {index} due to missing image URL.")
            continue

        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Du bist ein professioneller SEO-Experte für Alt-Texte. "
                            "Erstelle präzise, natürlich klingende Alt-Texte auf Deutsch, ideal für Suchmaschinenoptimierung (SEO). "
                            "Beschreibe den Bildinhalt klar und objektiv, mit Fokus auf sichtbare Objekte, Personen, Umgebung und Aktivitäten. "
                            "Begrenze den Alt-Text auf maximal 15 Wörter. "
                            "Verwende relevante Keywords und achte auf natürliche Sprache."
                        )
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    f"Bildanalyse:\n"
                                    f"- Dateiname: {filename}\n"
                                    f"- Vorhandener Alt-Text: {old_alt_text}\n"
                                    f"- Markenname: {store_name}\n\n"
                                    f"Aufgabe:\n"
                                    "- Beschreibe präzise, was auf dem Bild zu sehen ist (z. B. Objekt(e), Personen, Tätigkeit, Umgebung), inklusive relevanter Merkmale wie Farbe, Material oder Funktion.\n"
                                    f"- Baue den Markennamen ({store_name}) sinnvoll und flüssig in den Text ein, falls passend.\n"
                                    "- Verwende thematisch relevante SEO-Begriffe (z. B. Lederstuhl, Turnringe, Waldtraining), wenn möglich natürlich integriert.\n"
                                    "- Maximal 15 Wörter, vollständiger deutscher Satz ohne Listenstil."
                                )
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": image_url}
                            }
                        ]
                    }
                ],
                temperature=0.2,
                max_tokens=100
            )

            generated_text = response.choices[0].message.content
            df.at[index, "New Alt Text"] = generated_text.strip()
            print(f"✅ Generated new alt text for row {index} – {generated_text.strip()}")

            # 🔁 Save backup every 5 rows
            if index % backup_interval == 0:
                try:
                    with file_lock:
                        df.to_excel(temp_file, index=False)
                        print(f"💾 Temp backup written at row {index}")
                except Exception as e:
                    print(f"⚠️ Backup write failed at row {index}: {e}")

        except Exception as e:
            print(f"❌ Error generating alt text for row {index}: {str(e)}")
            continue

    # ✅ Final save with retry logic
    for attempt in range(3):
        try:
            with file_lock:
                df.to_excel(final_output_file, index=False)
            print(f"✅ All done! File saved as {final_output_file}")
            break
        except PermissionError:
            print(f"❌ Attempt {attempt + 1}: Permission denied. Retrying in 2s...")
            time.sleep(2)
        except Exception as e:
            print(f"❌ Unexpected error during final save: {e}")
            break
    else:
        print(f"❌ Final save failed after 3 attempts. Temp file saved at: {temp_file}")

def beautify_store_name(store_name):
    # Replace hyphens and underscores with spaces
    name = store_name.replace('-', ' ').replace('_', ' ')
    # Remove extra spaces (if any)
    name = ' '.join(name.split())
    # Capitalize first letter of each word
    name = name.title()
    return name

def check_thread(thread, callback):
    if thread.is_alive():
        # If the thread is still running, check again after 100 ms
        root.after(100, check_thread, thread, callback)
    else:
        # If the thread has finished, run the callback
        callback()

# Function to load the rest of the logic after GUI is created
def load_background_logic():
    print("Loading additional logic (e.g., imports)...")
    global pandas, requests, openpyxl, datetime, math, ET, base64
    
    import requests  # Example of delayed import
    from datetime import datetime
    import math
    import xml.etree.ElementTree as ET
    import base64
    print("Additional logic loaded successfully.")

    
# Set up Tkinter GUI
root = tk.Tk()
root.title("Shopify Tool")

# Let row 0 (text area) expand, and column 0+1 share space
root.rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=1)

# Text area in row 0
text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD)
text_area.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)

sys.stdout = RedirectOutput(text_area)

# Buttons in rows below
download_button = tk.Button(root, text="Download", command=start_download, width=20, height=2)
download_button.grid(row=1, column=0, pady=5)

upload_button = tk.Button(root, text="Upload", command=start_upload, width=20, height=2)
upload_button.grid(row=1, column=1, pady=5)

collection_download_button = tk.Button(root, text="Download Collections", command=start_collection_download, width=20, height=2)
collection_download_button.grid(row=2, column=0, pady=5)

collection_upload_button = tk.Button(root, text="Upload Collections", command=start_collection_upload, width=20, height=2)
collection_upload_button.grid(row=2, column=1, pady=5)

file_alt_download_button = tk.Button(root, text="Download Files Alt Texts", command=lambda: threading.Thread(target=download_shopify_files_alt_texts).start(), width=25, height=2)
file_alt_download_button.grid(row=3, column=0, pady=5)

file_alt_upload_button = tk.Button(root, text="Upload Files Alt Texts", command=lambda: threading.Thread(target=upload_shopify_files_alt_texts).start(), width=25, height=2)
file_alt_upload_button.grid(row=3, column=1, pady=5)

seo_alt_text_button = tk.Button(root, text="Generate SEO Alt Texts (AI)", command=lambda: threading.Thread(target=generate_seo_alt_texts).start(), width=30, height=2)
seo_alt_text_button.grid(row=4, column=0, columnspan=2, pady=5)

# Set window size
root.geometry("500x700")

# Show the window first, then load heavy logic in the background
root.after(100, lambda: threading.Thread(target=load_background_logic).start())

# Run the application
root.mainloop()
